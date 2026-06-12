"""Excel Dashboard Generator for Execution Decisions."""
import os
from datetime import datetime
from typing import Dict, Any

try:
    import openpyxl
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExecutionExcelGenerator:
    """
    Generate Excel Dashboard with multiple sheets.
    
    Sheets:
    1. Report Summary
    2. Live Market Data
    3. ICT Analysis
    4. DXY Analysis
    5. News Analysis
    6. Trade Levels
    7. Position Sizing
    8. Execution Decision
    9. Risk Analysis
    10. Final Summary
    """
    
    # Colors
    NAVY = 'FF0D1B2A'
    GOLD = 'FFC9A84C'
    GREEN = 'FF1E8449'
    RED = 'FFC0392B'
    ORANGE = 'FFE67E22'
    LTGRAY = 'FFF2F4F7'
    WHITE = 'FFFFFFFF'
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")
    
    def generate(self, data: Dict[str, Any], output_folder: str) -> str:
        """
        Generate the Excel dashboard.
        
        Args:
            data: Complete analysis data
            output_folder: Folder to save the Excel file
            
        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"Instant_Trade_Decision_{timestamp}.xlsx"
        filepath = os.path.join(output_folder, filename)
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, data)
        self._create_live_market_sheet(wb, data)
        self._create_ict_sheet(wb, data)
        self._create_dxy_sheet(wb, data)
        self._create_news_sheet(wb, data)
        self._create_trade_levels_sheet(wb, data)
        self._create_position_sizing_sheet(wb, data)
        self._create_decision_sheet(wb, data)
        self._create_risk_sheet(wb, data)
        self._create_final_summary_sheet(wb, data)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
    
    def _style_header(self, cell):
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _style_cell(self, cell, is_alt=False):
        """Apply standard cell styling."""
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if is_alt:
            cell.fill = PatternFill(start_color=self.LTGRAY, end_color=self.LTGRAY, fill_type='solid')
    
    def _add_header_row(self, ws, row, headers):
        """Add a header row to worksheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._style_header(cell)
    
    def _create_summary_sheet(self, wb, data: Dict):
        """Create Report Summary sheet."""
        ws = wb.create_sheet('01 Summary')
        
        # Title
        ws['A1'] = 'INSTANT TRADE DECISION - SUMMARY'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Basic info
        ws['A3'] = 'Pair:'
        ws['B3'] = data.get('pair', 'N/A')
        ws['A4'] = 'Direction:'
        ws['B4'] = data.get('direction', 'N/A')
        ws['A5'] = 'Timestamp:'
        ws['B5'] = data.get('timestamp', datetime.now().isoformat())
        
        # Decision
        decision = data.get('final_decision', {})
        ws['A7'] = 'DECISION:'
        ws['B7'] = decision.get('action', 'N/A')
        ws['B7'].font = Font(bold=True, size=14)
        
        # Color based on decision
        action = decision.get('action', '')
        if 'EXECUTE' in action:
            ws['B7'].fill = PatternFill(start_color=self.GREEN, end_color=self.GREEN, fill_type='solid')
            ws['B7'].font = Font(bold=True, size=14, color='FFFFFF')
        elif 'ABORT' in action:
            ws['B7'].fill = PatternFill(start_color=self.RED, end_color=self.RED, fill_type='solid')
            ws['B7'].font = Font(bold=True, size=14, color='FFFFFF')
        
        ws['A8'] = 'Confidence:'
        ws['B8'] = f"{decision.get('confidence', 0)}%"
        ws['A9'] = 'Instant Score:'
        ws['B9'] = f"{decision.get('instant_score', 0)}/28"
        ws['A10'] = 'Trade Quality:'
        ws['B10'] = f"{decision.get('trade_quality', 0)}/100"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_live_market_sheet(self, wb, data: Dict):
        """Create Live Market Data sheet."""
        ws = wb.create_sheet('02 Live Market')
        
        live_data = data.get('live_data', {})
        price_data = live_data.get('price_data', {})
        tech_data = live_data.get('technical_data', {})
        
        # Price data
        ws['A1'] = 'PRICE DATA'
        ws['A1'].font = Font(bold=True, size=12)
        
        headers = ['Metric', 'Value']
        self._add_header_row(ws, 2, headers)
        
        price_rows = [
            ['Current Price', price_data.get('current', 'N/A')],
            ['Bid', price_data.get('bid', 'N/A')],
            ['Ask', price_data.get('ask', 'N/A')],
            ['High', price_data.get('high', 'N/A')],
            ['Low', price_data.get('low', 'N/A')],
            ['Day Change %', f"{price_data.get('day_change_pct', 0)}%"],
            ['Week Change %', f"{price_data.get('week_change_pct', 0)}%"],
        ]
        
        for i, row in enumerate(price_rows, 3):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        # Technical data
        ws['A12'] = 'TECHNICAL INDICATORS'
        ws['A12'].font = Font(bold=True, size=12)
        
        self._add_header_row(ws, 13, ['Indicator', 'Value', 'Signal'])
        
        rsi = tech_data.get('rsi', {})
        macd = tech_data.get('macd', {})
        adx = tech_data.get('adx', {})
        
        tech_rows = [
            ['RSI (14)', rsi.get('value', 'N/A'), rsi.get('signal', 'N/A')],
            ['MACD', macd.get('histogram', 'N/A'), macd.get('trend', 'N/A')],
            ['ADX', adx.get('value', 'N/A'), adx.get('trend_strength', 'N/A')],
            ['Consensus', tech_data.get('consensus', 'N/A'), '-'],
        ]
        
        for i, row in enumerate(tech_rows, 14):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        # Adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_ict_sheet(self, wb, data: Dict):
        """Create ICT Analysis sheet."""
        ws = wb.create_sheet('03 ICT Analysis')
        
        ict = data.get('ict_analysis', {})
        
        ws['A1'] = 'ICT SMART MONEY ANALYSIS'
        ws['A1'].font = Font(bold=True, size=12)
        
        self._add_header_row(ws, 2, ['Component', 'Status', 'Details'])
        
        liquidity = ict.get('liquidity', {})
        ob = ict.get('order_blocks', {})
        fvg = ict.get('fair_value_gaps', {})
        zones = ict.get('zones', {})
        smt = ict.get('smt_divergence', {})
        
        rows = [
            ['Liquidity Sweep', 'Yes' if liquidity.get('sweep_detected') else 'No',
             liquidity.get('sweep_direction', 'None')],
            ['Order Block', 'Yes' if ob.get('active_ob') else 'No',
             ob.get('ob_direction', 'None')],
            ['Fair Value Gap', 'Yes' if fvg.get('fvg_present') else 'No',
             fvg.get('fvg_direction', 'None')],
            ['Zone', zones.get('zone', 'N/A'), f"{zones.get('zone_percentage', 0)}%"],
            ['SMT Divergence', 'Yes' if smt.get('detected') else 'No',
             smt.get('divergence_type', 'None')],
        ]
        
        for i, row in enumerate(rows, 3):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        # ICT Score
        alignment = ict.get('alignment_score', {})
        ws['A10'] = 'ICT Alignment Score:'
        ws['B10'] = f"{alignment.get('score', 0)}/100"
        ws['B10'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
    
    def _create_dxy_sheet(self, wb, data: Dict):
        """Create DXY Analysis sheet."""
        ws = wb.create_sheet('04 DXY Analysis')
        
        dxy = data.get('dxy_analysis', {})
        
        ws['A1'] = 'DXY ANALYSIS'
        ws['A1'].font = Font(bold=True, size=12)
        
        self._add_header_row(ws, 2, ['Metric', 'Value'])
        
        rows = [
            ['DXY Price', dxy.get('dxy_price', 'N/A')],
            ['DXY Change', f"{dxy.get('dxy_change', 0)}%"],
            ['DXY Trend', dxy.get('dxy_trend', 'N/A')],
            ['Classification', dxy.get('classification', {}).get('status', 'N/A')],
            ['Supports Trade', 'Yes' if dxy.get('supports_trade') else 'No'],
        ]
        
        for i, row in enumerate(rows, 3):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _create_news_sheet(self, wb, data: Dict):
        """Create News Analysis sheet."""
        ws = wb.create_sheet('05 News Analysis')
        
        news = data.get('news_gate', {})
        
        ws['A1'] = 'NEWS GATE ANALYSIS'
        ws['A1'].font = Font(bold=True, size=12)
        
        ws['A3'] = 'Risk Level:'
        ws['B3'] = news.get('risk_level', 'N/A')
        
        # Color based on risk
        risk = news.get('risk_level', '')
        if risk == 'BLOCKED':
            ws['B3'].fill = PatternFill(start_color=self.RED, end_color=self.RED, fill_type='solid')
            ws['B3'].font = Font(bold=True, color='FFFFFF')
        elif risk == 'CAUTION':
            ws['B3'].fill = PatternFill(start_color=self.ORANGE, end_color=self.ORANGE, fill_type='solid')
        elif risk == 'CLEAR':
            ws['B3'].fill = PatternFill(start_color=self.GREEN, end_color=self.GREEN, fill_type='solid')
            ws['B3'].font = Font(bold=True, color='FFFFFF')
        
        ws['A4'] = 'Recommendation:'
        ws['B4'] = news.get('recommendation', 'N/A')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
    
    def _create_trade_levels_sheet(self, wb, data: Dict):
        """Create Trade Levels sheet."""
        ws = wb.create_sheet('06 Trade Levels')
        
        levels = data.get('trade_levels', {})
        level_data = levels.get('levels', {}) if levels else {}
        
        ws['A1'] = 'TRADE LEVELS'
        ws['A1'].font = Font(bold=True, size=12)
        
        self._add_header_row(ws, 2, ['Level', 'Price', 'Pips', 'R:R'])
        
        rows = [
            ['Entry', level_data.get('entry', 'N/A'), '-', '-'],
            ['Stop Loss', level_data.get('stop_loss', 'N/A'),
             level_data.get('sl_pips', 'N/A'), '-'],
            ['TP1 (25%)', level_data.get('tp1', 'N/A'),
             level_data.get('tp1_pips', 'N/A'), '1:1.5'],
            ['TP2 (25%)', level_data.get('tp2', 'N/A'),
             level_data.get('tp2_pips', 'N/A'), '1:3'],
            ['TP3 (25%)', level_data.get('tp3', 'N/A'),
             level_data.get('tp3_pips', 'N/A'), '1:5'],
            ['TP4 (25%)', level_data.get('tp4', 'N/A'),
             level_data.get('tp4_pips', 'N/A'), '1:8'],
        ]
        
        for i, row in enumerate(rows, 3):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
    
    def _create_position_sizing_sheet(self, wb, data: Dict):
        """Create Position Sizing sheet."""
        ws = wb.create_sheet('07 Position Sizing')
        
        levels = data.get('trade_levels', {})
        sizing = levels.get('sizing', {}) if levels else {}
        
        ws['A1'] = 'POSITION SIZING'
        ws['A1'].font = Font(bold=True, size=12)
        
        self._add_header_row(ws, 2, ['Parameter', 'Value'])
        
        rows = [
            ['Account Size', f"${data.get('account_size', 'N/A')}"],
            ['Max Risk %', f"{data.get('max_risk', 'N/A')}%"],
            ['Leverage', f"{data.get('leverage', 'N/A')}:1"],
            ['Lot Size', sizing.get('lot_size', 'N/A')],
            ['Risk Amount', f"${sizing.get('risk_amount', 'N/A')}"],
            ['Max Loss', f"${sizing.get('max_loss', 'N/A')}"],
            ['Margin Required', f"${sizing.get('margin_required', 'N/A')}"],
        ]
        
        for i, row in enumerate(rows, 3):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _create_decision_sheet(self, wb, data: Dict):
        """Create Execution Decision sheet."""
        ws = wb.create_sheet('08 Decision')
        
        decision = data.get('final_decision', {})
        
        ws['A1'] = 'EXECUTION DECISION'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'DECISION:'
        ws['A3'].font = Font(bold=True, size=12)
        ws['B3'] = decision.get('action', 'N/A')
        ws['B3'].font = Font(bold=True, size=16)
        
        # Color based on decision
        action = decision.get('action', '')
        if 'EXECUTE' in action:
            ws['B3'].fill = PatternFill(start_color=self.GREEN, end_color=self.GREEN, fill_type='solid')
            ws['B3'].font = Font(bold=True, size=16, color='FFFFFF')
        elif 'WAIT' in action:
            ws['B3'].fill = PatternFill(start_color=self.GOLD, end_color=self.GOLD, fill_type='solid')
        elif 'MONITOR' in action:
            ws['B3'].fill = PatternFill(start_color=self.ORANGE, end_color=self.ORANGE, fill_type='solid')
        elif 'ABORT' in action:
            ws['B3'].fill = PatternFill(start_color=self.RED, end_color=self.RED, fill_type='solid')
            ws['B3'].font = Font(bold=True, size=16, color='FFFFFF')
        
        ws['A5'] = 'Reason:'
        ws['B5'] = decision.get('reason', 'N/A')
        
        ws['A6'] = 'Confidence:'
        ws['B6'] = f"{decision.get('confidence', 0)}%"
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
    
    def _create_risk_sheet(self, wb, data: Dict):
        """Create Risk Analysis sheet."""
        ws = wb.create_sheet('09 Risk Analysis')
        
        ws['A1'] = 'RISK ANALYSIS'
        ws['A1'].font = Font(bold=True, size=12)
        
        # Pro test results
        pro_test = data.get('pro_test', {})
        tests = pro_test.get('tests', {})
        
        self._add_header_row(ws, 2, ['Test', 'Result'])
        
        rows = [
            ['FTMO Approve', 'PASS' if tests.get('ftmo_approve') else 'FAIL'],
            ['Hedge Fund Enter', 'PASS' if tests.get('hedge_fund_enter') else 'FAIL'],
            ['Prop Trader Take', 'PASS' if tests.get('prop_trader_take') else 'FAIL'],
            ['Only Trade Today', 'PASS' if tests.get('only_trade_today') else 'FAIL'],
        ]
        
        for i, row in enumerate(rows, 3):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
                if j == 2:
                    if value == 'PASS':
                        cell.fill = PatternFill(start_color=self.GREEN, end_color=self.GREEN, fill_type='solid')
                        cell.font = Font(color='FFFFFF')
                    else:
                        cell.fill = PatternFill(start_color=self.RED, end_color=self.RED, fill_type='solid')
                        cell.font = Font(color='FFFFFF')
        
        ws['A9'] = 'Overall:'
        ws['B9'] = pro_test.get('verdict', 'N/A')
        ws['B9'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_final_summary_sheet(self, wb, data: Dict):
        """Create Final Summary sheet."""
        ws = wb.create_sheet('10 Final Summary')
        
        ws['A1'] = 'FINAL SUMMARY'
        ws['A1'].font = Font(bold=True, size=14)
        
        decision = data.get('final_decision', {})
        instant_score = data.get('instant_score', {})
        trade_quality = data.get('trade_quality', {})
        
        self._add_header_row(ws, 3, ['Metric', 'Value', 'Status'])
        
        rows = [
            ['Pair', data.get('pair', 'N/A'), '-'],
            ['Direction', data.get('direction', 'N/A'), '-'],
            ['Decision', decision.get('action', 'N/A'), decision.get('emoji', '')],
            ['Confidence', f"{decision.get('confidence', 0)}%",
             '✓' if decision.get('confidence', 0) >= 80 else '✗'],
            ['Instant Score', f"{instant_score.get('total', 0)}/28",
             '✓' if instant_score.get('total', 0) >= 14 else '✗'],
            ['Trade Quality', f"{trade_quality.get('total', 0)}/100",
             '✓' if trade_quality.get('total', 0) >= 70 else '✗'],
            ['Can Trade', 'Yes' if decision.get('can_trade') else 'No',
             '✓' if decision.get('can_trade') else '✗'],
        ]
        
        for i, row in enumerate(rows, 4):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                self._style_cell(cell, i % 2 == 0)
        
        # Timestamp
        ws['A13'] = 'Generated:'
        ws['B13'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
